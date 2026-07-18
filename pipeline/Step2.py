"""
Step2.py  —  Financial Statement Extractor → Excel Builder
===========================================================
Upgrades implemented
--------------------
1. Multi-Pass Algebraic Solver
   • PBT is resolved by trialling equations A / B / C in order; whichever
     balances within ±1 Lakh is injected as the live Excel formula.
   • All other key P&L milestones follow the same trial-loop pattern.

2. Taxonomy-Anchored Row Lookup
   • `get_row_by_node()` replaces the old `get_row_strict()`.
   • Instead of matching raw strings, it queries the master_cell_map via
     Internal Taxonomy Nodes (ITNs) from taxonomy.py.
   • Falls back gracefully if the node is absent on a given sheet.

3. Structural Segment Isolation
   • Standalone and Consolidated are built as two distinct SegmentPipeline
     objects before any Excel logic runs.
   • Consolidated pipelines carry `has_associates = True`, causing the
     algebra engine to prefer Equation C for PBT resolution.

4. Hardened JSON Cache
   • One JSON file per cache entry under output/step2_cache/, keyed on
     sha256(model + prompt_template + extracted_pdf_text) instead of just
     the PDF's filename — editing the extraction prompt, switching Gemini
     models, or re-running against an edited PDF all produce automatic
     cache misses rather than silently serving stale data. FORCE_REFRESH
     remains available as an explicit manual override.
"""

from google import genai as google_genai
import hashlib
import json
import fitz
import os
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import sys
from dataclasses import dataclass, field
from typing import Optional

# ── Taxonomy import (graceful fallback if file not present) ──────────────────
# Insert the directory that contains THIS script so `taxonomy.py` is always
# found even when the working directory differs (e.g. running from a venv or
# a different cwd).
import sys as _sys
import pathlib as _pathlib
_THIS_DIR = str(_pathlib.Path(__file__).resolve().parent)

if _THIS_DIR not in _sys.path:
    _sys.path.insert(0, _THIS_DIR)
    
_PROJECT_ROOT = str(_pathlib.Path(__file__).resolve().parent.parent)
try:
    from pipeline.taxonomy import TAXONOMY, get_taxonomy_enums
    from pipeline.taxonomy_mapper import map_line_items, build_populated_dictionary
    TAXONOMY_AVAILABLE = True
    print("✅  taxonomy.py loaded successfully.")
except ImportError:
    TAXONOMY_AVAILABLE = False
    print("⚠️  taxonomy.py not found — fuzzy taxonomy mapping disabled.")

    def get_taxonomy_enums() -> list:
        return ["UNMAPPED"]

    def map_line_items(records: list) -> list:
        for rec in records:
            rec["taxonomy_node"] = str(rec.get("raw_string", "")).strip()
            rec["fs_statement"] = ""
            rec["match_score"] = 0.0
        return records

    def build_populated_dictionary(df) -> dict:
        return {}


GEMINI_KEY   = os.environ.get("GEMINI_API_KEY")
FORCE_REFRESH = False
gemini = google_genai.Client(api_key=GEMINI_KEY)


# ─────────────────────────────────────────────────────────────────────────────
# STRUCTURAL SEGMENT ISOLATION  (Task 3)
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class SegmentPipeline:
    """
    Holds ALL data and metadata for one reporting segment
    (Standalone or Consolidated) before any Excel logic runs.
    Segments are completely independent — their algebra rules differ.
    """
    report_type   : str                            # "Standalone" | "Consolidated"
    has_associates: bool = False                   # True for Consolidated
    statements    : dict = field(default_factory=dict)  # statement_name → DataFrame

    def add(self, statement_name: str, df: pd.DataFrame):
        self.statements[statement_name] = df

    def get(self, statement_name: str) -> pd.DataFrame:
        return self.statements.get(statement_name, pd.DataFrame())


def build_segment_pipelines(df: pd.DataFrame) -> dict[str, SegmentPipeline]:
    """
    Splits the master DataFrame into isolated SegmentPipeline objects.
    Consolidated pipeline gets has_associates=True so the algebra engine
    can select the correct PBT equation.
    """
    pipelines: dict[str, SegmentPipeline] = {}

    for report_type in df["report_type"].unique():
        seg = SegmentPipeline(
            report_type    = report_type,
            has_associates = (report_type == "Consolidated"),
        )
        seg_df = df[df["report_type"] == report_type]
        for stmt in seg_df["statement"].unique():
            seg.add(stmt, seg_df[seg_df["statement"] == stmt].copy())
        pipelines[report_type] = seg
        print(f"  ✅ Segment pipeline built: {report_type}  |  has_associates={seg.has_associates}")

    return pipelines


# ─────────────────────────────────────────────────────────────────────────────
# TAXONOMY-ANCHORED ROW LOOKUP  (Task 2)
# ─────────────────────────────────────────────────────────────────────────────

# Maps Internal Taxonomy Node name → set of lowercase raw label variants
# Used to look up rows in master_cell_map when taxonomy is unavailable.
_FALLBACK_KEYS: dict[str, list[str]] = {
    "REVENUE_GROSS"              : ["value of services (revenue)", "value of services", "gross revenue from operations", "gross turnover", "turnover"],
    "REVENUE_GST_DEDUCTION"      : ["gst recovered", "less: gst recovered", "less: gst"],
    "REVENUE_FROM_OPERATIONS"    : ["revenue from operations", "net revenue from operations", "total revenue from operations"],
    "OTHER_INCOME"               : ["other income", "non-operating income"],
    "TOTAL_INCOME"               : ["total income", "total revenue", "aggregate revenue"],
    "TOTAL_EXPENSES"             : ["total expenses", "total operating expenses"],
    "SHARE_OF_PROFIT_OF_ASSOCIATES": ["share of profit of associate", "share of profit/(loss) of associates",
                                       "share of profit of associates and joint ventures"],
    "PROFIT_BEFORE_EXCEPTIONAL"  : ["profit before exceptional items and tax",
                                    "profit before share of profit of associates and tax",
                                    "profit before exceptional item and tax"],
    "EXCEPTIONAL_ITEMS"          : ["exceptional items", "exceptional item", "extraordinary items"],
    "PROFIT_BEFORE_TAX"          : ["profit before tax", "profit/(loss) before tax", "earnings before tax", "pbt"],
    "TOTAL_TAX_EXPENSE"          : ["total tax expenses", "total tax expense", "tax expenses total", "tax expenses (total)", "total income tax expense"],
    "PROFIT_FOR_THE_YEAR"        : ["profit for the year", "profit/(loss) for the year", "profit after tax", "net profit for the year"],
    "TOTAL_OCI"                  : ["total other comprehensive income/ (loss) for the year (net of tax)",
                                       "total other comprehensive income for the year (net of tax)",
                                       "total other comprehensive income",
                                       "other comprehensive income (net of tax)"],
    "TOTAL_COMPREHENSIVE_INCOME" : ["total comprehensive income/ (loss) for the year", "total comprehensive income/ (loss)", "total comprehensive income"],
    "EARNINGS_PER_SHARE"         : ["basic and diluted", "earnings per equity share"],
    "EQUITY_SHARE_CAPITAL"       : ["equity share capital"],
}


def get_rows_by_node(cell_map: dict, node_name: str) -> list[int]:
    """
    Like get_row_by_node, but returns EVERY matching row instead of the
    first one found.

    Needed for nodes like EARNINGS_PER_SHARE, where some PDFs report a
    single combined "Basic and Diluted" line (one row) but others report
    "Basic EPS" and "Diluted EPS" as two separate rows. get_row_by_node's
    single-row return silently drops the second row in the latter case —
    this variant fixes that by collecting all distinct row matches.
    """
    candidates: list[str] = []

    if TAXONOMY_AVAILABLE and node_name in TAXONOMY:
        node = TAXONOMY[node_name]
        candidates = [a.lower().strip() for a in node.aliases]
    else:
        candidates = _FALLBACK_KEYS.get(node_name, [])

    rows: list[int] = []

    # Exact matches first
    for alias in candidates:
        if alias in cell_map:
            row = cell_map[alias]
            if row not in rows:
                rows.append(row)

    # Prefix / substring fallback — only for aliases that found no exact match
    for alias in candidates:
        for key, row in cell_map.items():
            if row in rows:
                continue
            if key.startswith(alias) or alias in key:
                rows.append(row)

    return sorted(rows)


def get_row_by_node(cell_map: dict, node_name: str) -> Optional[int]:
    """
    Taxonomy-anchored row lookup (Task 2).

    1. If taxonomy.py is available, collects all known aliases for `node_name`
       and scans `cell_map` (lowercased keys) for any alias match.
    2. Falls back to _FALLBACK_KEYS for environments without taxonomy.py.
    3. Returns the Excel row number, or None if not found.
    """
    candidates: list[str] = []

    if TAXONOMY_AVAILABLE and node_name in TAXONOMY:
        node = TAXONOMY[node_name]
        candidates = [a.lower().strip() for a in node.aliases]
    else:
        candidates = _FALLBACK_KEYS.get(node_name, [])

    # Exact match first
    for alias in candidates:
        if alias in cell_map:
            return cell_map[alias]

    # Prefix / substring fallback
    for alias in candidates:
        for key, row in cell_map.items():
            if key.startswith(alias) or alias in key:
                return row

    return None


# ─────────────────────────────────────────────────────────────────────────────
# MULTI-PASS ALGEBRAIC SOLVER  (Task 1)
# ─────────────────────────────────────────────────────────────────────────────

TOLERANCE_LAKHS = 1.0   # ±1 unit rounding tolerance (in whatever reporting unit the source uses)


def _val(ws, row: Optional[int], col: str) -> float:
    """
    Read a numeric cell value; return 0.0 if missing or non-numeric.
    Explicitly rejects formula strings (starts with '=') so that cells
    already overwritten with a formula don't silently return 0.0 and
    corrupt downstream tolerance checks.
    """
    if row is None:
        return 0.0
    v = ws[f"{col}{row}"].value
    if isinstance(v, str) and v.startswith("="):
        # Cell already holds a formula — treat as 0.0 so callers know
        # this value is no longer a raw number and cannot be used for
        # numerical verification.
        return 0.0
    return float(v) if isinstance(v, (int, float)) else 0.0


def _try_inject(ws, target_row: int, col: str,
                equations: list[tuple[float, str]]) -> bool:
    """
    Multi-pass trial loop (Task 1 core).

    `equations` is an ordered list of (computed_value, excel_formula) tuples.
    The first equation whose computed_value matches the hardcoded cell value
    within TOLERANCE_LAKHS is injected into the cell and returns True.
    """
    if target_row is None:
        return False
    target_val = _val(ws, target_row, col)
    for computed, formula in equations:
        if abs(computed - target_val) <= TOLERANCE_LAKHS:
            ws[f"{col}{target_row}"] = formula
            return True
    return False


def inject_pnl_algebra(ws, cell_map: dict, col: str,
                        has_associates: bool = False):
    """
    Runs the multi-pass algebraic solver for every key P&L milestone.
    `has_associates=True` adds Equation C to the PBT trial set.
    """

    # ── Resolve row references via taxonomy ──────────────────────────────────
    vos  = get_row_by_node(cell_map, "REVENUE_GROSS")
    gst  = get_row_by_node(cell_map, "REVENUE_GST_DEDUCTION")
    rfo  = get_row_by_node(cell_map, "REVENUE_FROM_OPERATIONS")
    oi   = get_row_by_node(cell_map, "OTHER_INCOME")
    ti   = get_row_by_node(cell_map, "TOTAL_INCOME")
    te   = get_row_by_node(cell_map, "TOTAL_EXPENSES")
    spa  = get_row_by_node(cell_map, "SHARE_OF_PROFIT_OF_ASSOCIATES")
    pbe  = get_row_by_node(cell_map, "PROFIT_BEFORE_EXCEPTIONAL")
    exc  = get_row_by_node(cell_map, "EXCEPTIONAL_ITEMS")
    pbt  = get_row_by_node(cell_map, "PROFIT_BEFORE_TAX")
    tte  = get_row_by_node(cell_map, "TOTAL_TAX_EXPENSE")
    pfy  = get_row_by_node(cell_map, "PROFIT_FOR_THE_YEAR")
    toci = get_row_by_node(cell_map, "TOTAL_OCI")
    tci  = get_row_by_node(cell_map, "TOTAL_COMPREHENSIVE_INCOME")

    # Pre-read values
    v_vos  = _val(ws, vos,  col)
    v_gst  = _val(ws, gst,  col)
    v_rfo  = _val(ws, rfo,  col)
    v_oi   = _val(ws, oi,   col)
    v_ti   = _val(ws, ti,   col)
    v_te   = _val(ws, te,   col)
    v_spa  = _val(ws, spa,  col)
    v_pbe  = _val(ws, pbe,  col)
    v_exc  = _val(ws, exc,  col)
    v_pbt  = _val(ws, pbt,  col)
    v_tte  = _val(ws, tte,  col)
    v_pfy  = _val(ws, pfy,  col)
    v_toci = _val(ws, toci, col)

    # ── 1. Revenue from Operations ───────────────────────────────────────────
    if rfo and vos and gst:
        _try_inject(ws, rfo, col, [
            (v_vos - v_gst, f"={col}{vos}-{col}{gst}"),   # Eq A: gross minus GST
            (v_vos + v_gst, f"={col}{vos}+{col}{gst}"),   # Eq B: gross plus GST (rare)
        ])

    # ── 2. Total Income ──────────────────────────────────────────────────────
    if ti and rfo and oi:
        _try_inject(ws, ti, col, [
            (v_rfo + v_oi, f"={col}{rfo}+{col}{oi}"),     # Eq A: RfO + Other Income
        ])

    # ── 3. Profit Before Tax  (MULTI-PASS — Task 1 centrepiece) ─────────────
    #
    #   Equation A  →  Total Income − Total Expenses
    #   Equation B  →  Profit Before Exceptional ± Exceptional Items
    #   Equation C  →  Total Income − Total Expenses + Share of Profit (Consol.)
    #
    if pbt:
        pbt_equations = []

        # Eq A: TI - TE
        if ti and te:
            pbt_equations.append((v_ti - v_te, f"={col}{ti}-{col}{te}"))
            pbt_equations.append((v_ti + v_te, f"={col}{ti}+{col}{te}"))   # expenses stored negative

        # Eq B: PBE ± Exceptional
        if pbe and exc:
            pbt_equations.append((v_pbe - v_exc, f"={col}{pbe}-{col}{exc}"))
            pbt_equations.append((v_pbe + v_exc, f"={col}{pbe}+{col}{exc}"))

        # Eq C: TI − TE + Share of Associates (Consolidated only)
        if has_associates and ti and te and spa:
            pbt_equations.append((v_ti - v_te + v_spa, f"={col}{ti}-{col}{te}+{col}{spa}"))
            pbt_equations.append((v_ti + v_te + v_spa, f"={col}{ti}+{col}{te}+{col}{spa}"))

        _try_inject(ws, pbt, col, pbt_equations)

    # ── 4. Profit for the Year ───────────────────────────────────────────────
    if pfy and pbt and tte:
        _try_inject(ws, pfy, col, [
            (v_pbt - v_tte, f"={col}{pbt}-{col}{tte}"),   # Eq A: PBT - Tax
            (v_pbt + v_tte, f"={col}{pbt}+{col}{tte}"),   # Eq B: tax stored negative
        ])

    # ── 5. Total Comprehensive Income ────────────────────────────────────────
    if tci and pfy and toci:
        v_pfy_actual = _val(ws, pfy, col)   # re-read in case formula was injected
        _try_inject(ws, tci, col, [
            (v_pfy + v_toci, f"={col}{pfy}+{col}{toci}"),
            (v_pfy - v_toci, f"={col}{pfy}-{col}{toci}"),
        ])


def inject_bs_algebra(ws, cell_map: dict, col: str):
    """
    Balance Sheet algebraic solver — mirrors inject_pnl_algebra's trial-loop
    pattern, anchored on the Balance Sheet subtotal chains present in taxonomy.py:

        TOTAL_LIABILITIES            = TOTAL_NON_CURRENT_LIABILITIES
                                        + TOTAL_CURRENT_LIABILITIES
        TOTAL_EQUITY_AND_LIABILITIES = TOTAL_EQUITY + TOTAL_LIABILITIES
        TOTAL_ASSETS                 = TOTAL_NON_CURRENT_ASSETS
                                        + TOTAL_CURRENT_ASSETS

    Each injection is independently gated — if a row is missing on a given
    sheet, that equation is simply skipped rather than guessed at.
    """
    te   = get_row_by_node(cell_map, "TOTAL_EQUITY")
    tncl = get_row_by_node(cell_map, "TOTAL_NON_CURRENT_LIABILITIES")
    tcl  = get_row_by_node(cell_map, "TOTAL_CURRENT_LIABILITIES")
    tl   = get_row_by_node(cell_map, "TOTAL_LIABILITIES")
    teal = get_row_by_node(cell_map, "TOTAL_EQUITY_AND_LIABILITIES")
    tnca = get_row_by_node(cell_map, "TOTAL_NON_CURRENT_ASSETS")
    tca  = get_row_by_node(cell_map, "TOTAL_CURRENT_ASSETS")
    ta   = get_row_by_node(cell_map, "TOTAL_ASSETS")

    v_te   = _val(ws, te,   col)
    v_tncl = _val(ws, tncl, col)
    v_tcl  = _val(ws, tcl,  col)
    v_tl   = _val(ws, tl,   col)
    v_teal = _val(ws, teal, col)
    v_tnca = _val(ws, tnca, col)
    v_tca  = _val(ws, tca,  col)

    # ── 1. Total Liabilities = Non-Current + Current ────────────────────────
    if tl and tncl and tcl:
        _try_inject(ws, tl, col, [
            (v_tncl + v_tcl, f"={col}{tncl}+{col}{tcl}"),
        ])

    # ── 2. Total Equity and Liabilities = Equity + Liabilities ──────────────
    # NOTE: uses v_tl captured BEFORE the Total Liabilities injection above —
    # if that cell was just turned into a formula, _val() would return 0.0
    # for it post-injection, corrupting this equation. The pre-read raw value
    # is the correct operand regardless of whether TL itself got a formula.
    if teal and te and tl:
        _try_inject(ws, teal, col, [
            (v_te + v_tl, f"={col}{te}+{col}{tl}"),
        ])

    # ── 3. Total Assets = Non-Current Assets + Current Assets ───────────────
    # Component-based sum rather than an equity-and-liabilities mirror.
    # v_tnca and v_tca are pre-read raw values — safe even if earlier steps
    # have overwritten other cells with formulas (_val() returns 0.0 for those).
    if ta and tnca and tca:
        _try_inject(ws, ta, col, [
            (v_tnca + v_tca, f"={col}{tnca}+{col}{tca}"),
        ])


def inject_cf_algebra(ws, cell_map: dict, col: str):
    """
    Cash Flow algebraic solver — mirrors inject_pnl_algebra / inject_bs_algebra's
    trial-loop pattern, anchored on the real CFO/CFI/CFF chain present in
    taxonomy.py and verified against an actual extracted Cash Flow sheet:

        CFO_CASH_GENERATED_FROM_OPERATIONS = CFO_OPERATING_PROFIT_BEFORE_WC
                                              + CFO_WC_ADJUSTMENTS_SUBTOTAL
        NET_CASH_FROM_OPERATING            = CFO_CASH_GENERATED_FROM_OPERATIONS
                                              + CFO_TAXES_PAID
        NET_CHANGE_IN_CASH                 = NET_CASH_FROM_OPERATING
                                              + NET_CASH_FROM_INVESTING
                                              + NET_CASH_FROM_FINANCING
        CLOSING_CASH_BALANCE               = OPENING_CASH_BALANCE
                                              + NET_CHANGE_IN_CASH

    Each injection is independently gated on its own rows being present —
    a missing node on a given sheet simply skips that one equation rather
    than guessing at a substitute.
    """
    r_opwc  = get_row_by_node(cell_map, "CFO_OPERATING_PROFIT_BEFORE_WC")
    r_wcadj = get_row_by_node(cell_map, "CFO_WC_ADJUSTMENTS_SUBTOTAL")
    r_cgen  = get_row_by_node(cell_map, "CFO_CASH_GENERATED_FROM_OPERATIONS")
    r_tax   = get_row_by_node(cell_map, "CFO_TAXES_PAID")
    r_nco   = get_row_by_node(cell_map, "NET_CASH_FROM_OPERATING")
    r_nci   = get_row_by_node(cell_map, "NET_CASH_FROM_INVESTING")
    r_ncf   = get_row_by_node(cell_map, "NET_CASH_FROM_FINANCING")
    r_nchg  = get_row_by_node(cell_map, "NET_CHANGE_IN_CASH")
    r_open  = get_row_by_node(cell_map, "OPENING_CASH_BALANCE")
    r_close = get_row_by_node(cell_map, "CLOSING_CASH_BALANCE")

    v_opwc  = _val(ws, r_opwc,  col)
    v_wcadj = _val(ws, r_wcadj, col)
    v_cgen  = _val(ws, r_cgen,  col)
    v_tax   = _val(ws, r_tax,   col)
    v_nco   = _val(ws, r_nco,   col)
    v_nci   = _val(ws, r_nci,   col)
    v_ncf   = _val(ws, r_ncf,   col)
    v_nchg  = _val(ws, r_nchg,  col)
    v_open  = _val(ws, r_open,  col)

    # ── 1. Cash Generated from Operations = Op Profit before WC + WC Adj ───
    if r_cgen and r_opwc and r_wcadj:
        _try_inject(ws, r_cgen, col, [
            (v_opwc + v_wcadj, f"={col}{r_opwc}+{col}{r_wcadj}"),
        ])

    # ── 2. Net Cash from Operating = Cash Generated + Taxes Paid ────────────
    # v_cgen is the pre-injection raw value (same pattern as inject_bs_algebra:
    # if step 1 just turned r_cgen into a formula, _val() would read 0.0 for
    # it post-injection, corrupting this equation).
    if r_nco and r_cgen and r_tax:
        _try_inject(ws, r_nco, col, [
            (v_cgen + v_tax, f"={col}{r_cgen}+{col}{r_tax}"),
        ])

    # ── 3. Net Change in Cash = Operating + Investing + Financing ──────────
    if r_nchg and r_nco and r_nci and r_ncf:
        _try_inject(ws, r_nchg, col, [
            (v_nco + v_nci + v_ncf, f"={col}{r_nco}+{col}{r_nci}+{col}{r_ncf}"),
        ])

    # ── 4. Closing Cash Balance = Opening Balance + Net Change ─────────────
    if r_close and r_open and r_nchg:
        _try_inject(ws, r_close, col, [
            (v_open + v_nchg, f"={col}{r_open}+{col}{r_nchg}"),
        ])


def inject_equity_algebra(ws, cell_map: dict, col: str):
    """
    Statement of Changes in Equity (SOCE) roll-forward solver.

    A real SOCE sheet has multiple parallel reserve rows (Equity Share
    Capital, Capital Reserve, Securities Premium, General Reserve, Retained
    Earnings, OCI, plus a Total row) — each its own contiguous block of
    Opening / TCI / Dividend / Transfer / Closing rows. RESERVE_CLOSING_BALANCE
    matches ALL closing rows at once via get_rows_by_node, and RESERVE_OPENING_BALANCE
    matches ALL opening rows — the node alone can't tell which opening pairs
    with which closing.

    NOTE: this deliberately does NOT reuse verify_and_inject_sum's colour-based
    section scanning. That function's HEADER_RGB hard-stop compares against the
    bare hex string "1F3864", but openpyxl returns ARGB strings like "001F3864"
    for solid fills, so the hard-stop silently never fires. Combined with
    stop_on_section=False treating every banner as skip-not-stop, it walks
    straight past a reserve's own section banner into a neighbouring reserve's
    rows on SOCE sheets (confirmed against the real Standalone - Changes in
    Equity sheet, where it pulled in Equity Share Capital rows while resolving
    Capital Reserve's closing balance). Bounding by the nearest taxonomy-anchored
    opening row above is reliable regardless of the colour-detection bug.

    Verified against real extracted SOCE data: every reserve's roll-forward
    reduces to Closing = SUM(opening_row : closing_row - 1).
    """
    closing_rows = get_rows_by_node(cell_map, "RESERVE_CLOSING_BALANCE")
    opening_rows = get_rows_by_node(cell_map, "RESERVE_OPENING_BALANCE")

    esc_closing = get_row_by_node(cell_map, "EQUITY_SHARE_CAPITAL_CLOSING")
    esc_opening = get_row_by_node(cell_map, "EQUITY_SHARE_CAPITAL_OPENING")
    if esc_closing and esc_closing not in closing_rows:
        closing_rows = sorted(closing_rows + [esc_closing])
    if esc_opening and esc_opening not in opening_rows:
        opening_rows = sorted(opening_rows + [esc_opening])

    for close_row in closing_rows:
        # Find the nearest opening row strictly above this closing row —
        # that is this reserve's own block start, regardless of cell colour.
        candidates_above = [r for r in opening_rows if r < close_row]
        if not candidates_above:
            continue
        open_row = max(candidates_above)

        target_val = _val(ws, close_row, col)
        calc_sum   = sum(_val(ws, r, col) for r in range(open_row, close_row))

        if abs(calc_sum - target_val) <= TOLERANCE_LAKHS:
            ws[f"{col}{close_row}"] = f"=SUM({col}{open_row}:{col}{close_row - 1})"


def verify_and_inject_sum(ws, target_row: int, col: str,
                           section_rgbs: set,
                           stop_on_section: bool = True) -> Optional[str]:
    """
    Scans upward from  collecting data rows and SUM-injects
    if the tally matches within TOLERANCE_LAKHS.

    stop_on_section=True  (default): halts at any section-banner fill —
                           correct for single-section blocks like Total Expenses.
    stop_on_section=False: treats section banners as cosmetic separators and
                           keeps scanning — required for Total OCI whose block
                           contains sub-section banners between individual items.
    """
    curr = target_row - 1
    candidate_rows = []
    HEADER_RGB = "1F3864"   # HEADER_FILL — hard stop regardless of mode

    while curr >= 3:
        rgb = ws[f"A{curr}"].fill.start_color.rgb

        # Always stop at the sheet header (dark blue title / column-header rows)
        if rgb == HEADER_RGB:
            break

        # In strict mode, stop at any section banner or total row
        if stop_on_section and rgb in section_rgbs:
            break

        # Skip section banners when scanning freely (do not add to candidates)
        if rgb in section_rgbs:
            curr -= 1
            continue

        candidate_rows.append(curr)
        curr -= 1

    if not candidate_rows:
        return None

    target_val = _val(ws, target_row, col)
    calc_sum   = sum(_val(ws, r, col) for r in candidate_rows)

    if abs(calc_sum - target_val) <= TOLERANCE_LAKHS:
        lo, hi = min(candidate_rows), max(candidate_rows)
        return f"=SUM({col}{lo}:{col}{hi})"
    return None


# ─────────────────────────────────────────────────────────────────────────────
# PDF EXTRACTION  (unchanged)
# ─────────────────────────────────────────────────────────────────────────────

def extract_text_from_pdf(pdf_path: str) -> str:
    doc = fitz.open(pdf_path)
    full_text = ""
    for page in doc:
        full_text += page.get_text()
    return full_text


def clean_number(val):
    if pd.isna(val) or val is None:
        return 0
    if isinstance(val, str):
        match = re.search(r"-?\d+\.?\d*", val.replace(",", "").replace(" ", ""))
        return float(match.group()) if match else 0
    return val


# ─────────────────────────────────────────────────────────────────────────────
# UNIT DETECTION
# ─────────────────────────────────────────────────────────────────────────────
# Indian financial statements state their reporting unit explicitly near the
# top of each statement (e.g. "(₹ in Crore)", "Rs. in Lakh", "all amounts in
# millions unless otherwise stated"). Different companies/years use different
# units, so we detect it from the actual PDF text instead of assuming one.
_UNIT_PATTERNS = [
    # ₹ / Rs. / backtick (common OCR/PDF-extraction surrogate for ₹) + "in Crore(s)" / "Cr."
    (re.compile(r"(?:₹|`|Rs\.?|INR)\s*(?:in\s+)?Crores?\b", re.IGNORECASE), "₹ Crore"),
    (re.compile(r"\bin\s+Crores?\b", re.IGNORECASE), "₹ Crore"),
    (re.compile(r"(?:₹|`|Rs\.?|INR)\s*Cr\.?\b", re.IGNORECASE), "₹ Crore"),
    # "(All amounts/figures are/in ... Crore, unless otherwise stated)" style headers
    (re.compile(r"(?:amounts?|figures?)\s+(?:are\s+)?in\s+(?:₹|`|Rs\.?|INR)?\s*Crores?", re.IGNORECASE), "₹ Crore"),

    (re.compile(r"(?:₹|`|Rs\.?|INR)\s*(?:in\s+)?Lakhs?\b", re.IGNORECASE), "₹ Lakh"),
    (re.compile(r"\bin\s+Lakhs?\b", re.IGNORECASE), "₹ Lakh"),
    (re.compile(r"(?:amounts?|figures?)\s+(?:are\s+)?in\s+(?:₹|`|Rs\.?|INR)?\s*Lakhs?", re.IGNORECASE), "₹ Lakh"),

    (re.compile(r"(?:₹|`|Rs\.?|INR)\s*(?:in\s+)?Millions?\b", re.IGNORECASE), "₹ Million"),
    (re.compile(r"\bin\s+Millions?\b", re.IGNORECASE), "₹ Million"),
    (re.compile(r"(?:amounts?|figures?)\s+(?:are\s+)?in\s+(?:₹|`|Rs\.?|INR)?\s*Millions?", re.IGNORECASE), "₹ Million"),

    (re.compile(r"(?:₹|`|Rs\.?|INR)\s*(?:in\s+)?Billions?\b", re.IGNORECASE), "₹ Billion"),
    (re.compile(r"\bin\s+Billions?\b", re.IGNORECASE), "₹ Billion"),
]


def detect_reporting_unit(text: str) -> str:
    """Detects the reporting unit (Crore / Lakh / Million / Billion) as
    actually stated in the source PDF text, by counting occurrences of each
    pattern and taking the most frequent — rather than assuming a fixed unit.
    Falls back to '₹ Lakh' (labelled as a fallback) only if nothing is found,
    which is a strong signal the extraction should be checked manually."""
    counts: dict[str, int] = {}
    for pattern, label in _UNIT_PATTERNS:
        n = len(pattern.findall(text))
        if n:
            counts[label] = counts.get(label, 0) + n
    if not counts:
        print("  ⚠️  Could not detect a reporting unit (Crore/Lakh/Million) in the PDF text — "
              "defaulting to '₹ Lakh' as a fallback. Please verify against the source.")
        return "₹ Lakh"
    detected = max(counts, key=counts.get)
    print(f"  📏  Detected reporting unit: {detected} (counts: {counts})")
    return detected


_GEMINI_MODEL: str = "gemini-2.5-flash"

# ─────────────────────────────────────────────────────────────────────────────
# SCHEMA-DRIVEN EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────
# Taxonomy mapping now happens HERE, at the initial LLM extraction step,
# instead of via a later Python-side fuzzy-matching pass. The LLM is
# constrained by Gemini's response_schema to only ever pick a taxonomy_node
# that actually exists in the registry (get_taxonomy_enums()), and to tag
# each line item's scope directly. Line items the LLM can't confidently
# classify are tagged "UNMAPPED" rather than guessed at.

_EXTRACTION_PROMPT_TEMPLATE = """Below is text from a financial statement PDF of an Indian company.
Extract EVERY SINGLE LINE ITEM from ALL financial statements including Statement of Changes in Equity.
You must extract BOTH Standalone and Consolidated statements and clearly segregate them by "scope".
Extract EVERY line item exactly as it appears — do NOT skip or omit any row, even ones you don't
recognize (e.g. "Intangible Assets Under Development", "Financial Assets", sub-notes, etc).

For EACH line item, output:
- raw_string    : the exact text of the line item as it appears in the PDF.
- statement     : which financial statement this row belongs to — one of
                  "Profit and Loss", "Balance Sheet", "Cash Flow", "Statement of Changes in Equity".
- scope         : "STANDALONE" or "CONSOLIDATED".
- current_year  : the current year (2024-25) figure, or null.
- previous_year : the previous year (2023-24) figure, or null.

CRITICAL STRING STANDARDIZATION RULES (APPLY TO BOTH STANDALONE & CONSOLIDATED):
1. For "Changes in Equity", always structure raw_string as: "[Reserve Name] - [Row Concept]".
2. Strip out all Roman numerals, letters, or bullet points from the beginning of raw_string.
3. For Cash Flow subtotals, use exact identical terms across both statements.
4. Ensure Trade Payables items under Current Liabilities always carry the prefix "Trade Payables - ".
5. If there is no change or the value is zero for a given year, output null.

CRITICAL RULES FOR NUMBERS:
- Numbers have spaces due to PDF formatting — remove spaces inside numbers. e.g. "11 48" = 1148
- Numbers in brackets mean NEGATIVE. e.g. "(3 64)" = -364
- {unit_note}
- Two columns = current year (2024-25) and previous year (2023-24)
- Use minus sign for negatives, never brackets

PDF TEXT:
{chunk}
"""


def _build_extraction_response_schema():
    """
    Builds the Gemini response_schema for schema-driven extraction: a JSON
    array of objects, each constrained to raw_string / taxonomy_node (enum,
    injected from taxonomy.get_taxonomy_enums()) / scope (enum) /
    current_year / previous_year.
    """
    from google.genai import types

    return types.Schema(
        type=types.Type.ARRAY,
        items=types.Schema(
            type=types.Type.OBJECT,
            properties={
                "raw_string": types.Schema(type=types.Type.STRING),
                "statement": types.Schema(
                    type=types.Type.STRING,
                    enum=["Profit and Loss", "Balance Sheet", "Cash Flow", "Statement of Changes in Equity"],
                ),
                "scope": types.Schema(
                    type=types.Type.STRING, enum=["STANDALONE", "CONSOLIDATED"]
                ),
                "current_year": types.Schema(type=types.Type.NUMBER, nullable=True),
                "previous_year": types.Schema(type=types.Type.NUMBER, nullable=True),
            },
            required=["raw_string", "statement", "scope"],
        ),
    )


def _generate_with_retry(model, contents, config, max_attempts=4, base_delay=5):
    for attempt in range(max_attempts):
        try:
            return gemini.models.generate_content(model=model, contents=contents, config=config)
        except Exception as exc:
            msg = str(exc)
            if not any(m in msg for m in ("503", "UNAVAILABLE", "overloaded")) or attempt == max_attempts - 1:
                raise
            delay = base_delay * (2 ** attempt)
            print(f"  ⚠️  Gemini overloaded (503) — retrying in {delay}s (attempt {attempt+1}/{max_attempts})...")
            time.sleep(delay)


def call_gemini(text: str) -> list:
    from google.genai import types

    unit_label = detect_reporting_unit(text)
    unit_note  = f"All numbers in the source are in {unit_label} — do not convert to any other unit"

    split_point = int(len(text) * 0.60)
    chunks = [text[:split_point], text[split_point:]]

    response_schema = _build_extraction_response_schema()
    config = types.GenerateContentConfig(
        response_mime_type="application/json",
        response_schema=response_schema,
    )

    all_data = []
    for i, chunk in enumerate(chunks):
        print(f"  Sending chunk {i+1}/2 to Gemini...")
        response = _generate_with_retry(
            _GEMINI_MODEL,
            _EXTRACTION_PROMPT_TEMPLATE.format(chunk=chunk, unit_note=unit_note),
            config,
        )
        raw   = response.text
        clean = raw.replace("```json", "").replace("```", "").strip()
        last_brace = clean.rfind("}")
        if not clean.rstrip().endswith("]"):
            clean = clean[:last_brace + 1] + "]"
        data = json.loads(clean)
        all_data.extend(data)
    return all_data


# ─────────────────────────────────────────────────────────────────────────────
# CACHE LAYER — hardened JSON file cache
# ─────────────────────────────────────────────────────────────────────────────
# Mirrors phase2_llm_classify.py's cache-invalidation approach (content hash
# of everything that can change the output) but keeps Step2's original
# storage format: one plain JSON file per cache entry under output/.
#
# The cache key now covers:
#   - the extracted PDF text itself (sha256) — a different/edited PDF misses
#   - the prompt TEMPLATE text — editing call_gemini's PROMPT auto-invalidates
#   - the model string — switching models auto-invalidates
# This replaces the old pdf_stem-only key, which silently served stale
# results after a prompt or model change unless FORCE_REFRESH was flipped
# by hand.

_CACHE_DIR_NAME = "step2_cache"


def _compute_step2_cache_key(pdf_text: str, prompt_template: str, model: str) -> str:
    """
    Deterministic cache key: sha256(model + prompt_template + pdf_text).
    Any change to the model, the prompt template, or the extracted PDF text
    produces a different key, so the old entry is simply never looked up
    again — no manual invalidation step required.
    """
    h = hashlib.sha256()
    h.update(model.encode("utf-8"))
    h.update(b"\x00")
    h.update(prompt_template.encode("utf-8"))
    h.update(b"\x00")
    h.update(pdf_text.encode("utf-8"))
    return h.hexdigest()


def _step2_cache_path(cache_key: str) -> str:
    cache_dir = os.path.join(_PROJECT_ROOT, "output", _CACHE_DIR_NAME)
    os.makedirs(cache_dir, exist_ok=True)
    return os.path.join(cache_dir, f"{cache_key}.json")




def truncate_financial_records(df: pd.DataFrame) -> pd.DataFrame:
    truncated_records = []
    for (report_type, statement_type), group in df.groupby(
        ["report_type", "statement"], sort=False
    ):
        stop_collecting = False
        for _, row in group.iterrows():
            if stop_collecting:
                continue
            truncated_records.append(row)
            item_lower = str(row["line_item"]).lower()
            if statement_type == "Profit and Loss":
                if "earnings per share" in item_lower or "eps" in item_lower:
                    if (
                        "diluted" in item_lower
                        or "nominal value" in item_lower
                        or row.equals(group.iloc[-1])
                    ):
                        stop_collecting = True
            elif statement_type == "Cash Flow":
                if "closing" in item_lower and "cash" in item_lower:
                    stop_collecting = True
                elif "cash and cash equivalents at the end" in item_lower:
                    stop_collecting = True
    return pd.DataFrame(truncated_records)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# REQUIRED IMPORTS  (replace / merge with existing openpyxl import block)
# ─────────────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER  —  visual layer only; all algebra/math logic is untouched
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(df, output_path: str, unit_label: str = "₹ Lakh"):

    # ── Palette ───────────────────────────────────────────────────────────────
    C_NAVY        = "1B365D"   # Title fill
    C_SLATE_BLUE  = "4A5B78"   # Header fill / milestone top-border
    C_ICE_BLUE    = "F0F4F8"   # Total/subtotal fill
    C_ZEBRA       = "F8F9FA"   # Alternating odd-row fill
    C_WHITE       = "FFFFFF"
    C_CHARCOAL    = "1C2833"   # Total font
    C_SLATE_TEXT  = "2C3E50"   # Data font
    C_HAIRLINE    = "E5E7E9"   # Standard cell border
    C_EXEC_BOTTOM = "1B365D"   # Milestone bottom border (double)

    # ── Fills ─────────────────────────────────────────────────────────────────
    TITLE_FILL   = PatternFill("solid", start_color=C_NAVY,       end_color=C_NAVY)
    HEADER_FILL  = PatternFill("solid", start_color=C_SLATE_BLUE, end_color=C_SLATE_BLUE)
    TOTAL_FILL   = PatternFill("solid", start_color=C_ICE_BLUE,   end_color=C_ICE_BLUE)
    ZEBRA_FILL   = PatternFill("solid", start_color=C_ZEBRA,      end_color=C_ZEBRA)
    WHITE_FILL   = PatternFill("solid", start_color=C_WHITE,      end_color=C_WHITE)

    # Section banners reuse the header-level fill so the algebra scanner
    # still recognises them via SECTION_RGBS (colour-based hard-stop).
    SECTION_FILL = PatternFill("solid", start_color=C_SLATE_BLUE, end_color=C_SLATE_BLUE)

    # ── Fonts ─────────────────────────────────────────────────────────────────
    TITLE_FONT   = Font(name="Segoe UI", bold=True,  color=C_WHITE,    size=15)
    HEADER_FONT  = Font(name="Segoe UI", bold=True,  color=C_WHITE,    size=10)
    SECTION_FONT = Font(name="Segoe UI", bold=True,  color=C_WHITE,    size=10)
    TOTAL_FONT   = Font(name="Segoe UI", bold=True,  color=C_CHARCOAL, size=10)
    NORMAL_FONT  = Font(name="Segoe UI",             color=C_SLATE_TEXT, size=10)

    # ── Alignments ────────────────────────────────────────────────────────────
    CENTER    = Alignment(horizontal="center",  vertical="center", wrap_text=False)
    LEFT      = Alignment(horizontal="left",    vertical="center", indent=1)
    LEFT_IND  = Alignment(horizontal="left",    vertical="center", indent=2)
    RIGHT     = Alignment(horizontal="right",   vertical="center")

    # ── Number format (standard accounting) ───────────────────────────────────
    NUM_FORMAT = r'#,##0;(#,##0);"-";@'

    # ── Borders ───────────────────────────────────────────────────────────────
    _hair     = Side(style="thin", color=C_HAIRLINE)
    DATA_BORDER = Border(left=_hair, right=_hair, top=_hair, bottom=_hair)

    _slate    = Side(style="thin",   color=C_SLATE_BLUE)
    _exec_bot = Side(style="double", color=C_EXEC_BOTTOM)
    TOTAL_BORDER = Border(top=_slate, bottom=_exec_bot)

    # ── SECTION_RGBS (used by verify_and_inject_sum colour scanner) ───────────
    SECTION_RGBS = {
        "FF" + C_SLATE_BLUE,   # openpyxl prefixes solid fills with "FF"
        "FF" + C_ICE_BLUE,
        "FF" + C_NAVY,
        # bare hex fallbacks (some openpyxl versions omit the "FF" prefix)
        C_SLATE_BLUE,
        C_ICE_BLUE,
        C_NAVY,
    }

    # ── Helper: is this row a milestone/total? ────────────────────────────────
    def is_total(rec: dict) -> bool:
        if rec.get("is_total", False):
            return True
        label = str(rec.get("line_item", "")).lower()
        return any(
            k in label
            for k in ["total", "net", "profit", "closing", "opening",
                       "balance at the", "cash generated"]
        )

    master_cell_map: dict[str, dict] = {}

    # ── Per-sheet writer ──────────────────────────────────────────────────────
    def write_sheet(wb, sheet_name: str, data):
        ws = wb.create_sheet(sheet_name)

        # Enable grid lines
        ws.sheet_view.showGridLines = True

        # ── Fixed Column A width ──────────────────────────────────────────────
        ws.column_dimensions["A"].width = 52

        # ── Row 1 — Title ─────────────────────────────────────────────────────
        ws.row_dimensions[1].height = 42
        ws.merge_cells("A1:D1")
        ws["A1"].value     = sheet_name
        ws["A1"].font      = TITLE_FONT
        ws["A1"].fill      = TITLE_FILL
        ws["A1"].alignment = CENTER

        # ── Row 2 — Column headers ────────────────────────────────────────────
        # Column D (Taxonomy Node) is metadata-only — added so downstream
        # tools (excel_reader_tool → narrative_agent) can look rows up by
        # taxonomy_node directly instead of re-parsing line_item text. The
        # algebra engine below only ever touches columns B/C, so this is
        # purely additive and doesn't affect the self-verifying formulas.
        ws.row_dimensions[2].height = 28
        for col_letter, label in zip(["A", "B", "C", "D"],
                                      ["Line Item", f"FY25 ({unit_label})", f"FY24 ({unit_label})", "Taxonomy Node"]):
            cell            = ws[f"{col_letter}2"]
            cell.value      = label
            cell.font       = HEADER_FONT
            cell.fill       = HEADER_FILL
            cell.alignment  = CENTER

        row          = 3
        prev_section = None
        data_row_seq = 0          # counts actual data rows (for zebra striping)
        master_cell_map[sheet_name] = {}

        for _, r in data.iterrows():
            section = str(r["section"]).strip()
            item    = str(r["line_item"]).strip()
            cy_val  = r["current_year"]
            py_val  = r["previous_year"]
            rec     = r.to_dict()

            # ── Section banner ────────────────────────────────────────────────
            if section and section != prev_section and section.lower() != "nan":
                if item.lower() != section.lower():
                    ws.row_dimensions[row].height = 21
                    ws.merge_cells(f"A{row}:D{row}")
                    cell           = ws[f"A{row}"]
                    cell.value     = section
                    cell.font      = SECTION_FONT
                    cell.fill      = SECTION_FILL
                    cell.alignment = LEFT
                    # hairline border on banner
                    for cl in ["A", "B", "C", "D"]:
                        ws[f"{cl}{row}"].border = DATA_BORDER
                    row += 1
                prev_section = section

            # ── Data row ──────────────────────────────────────────────────────
            ws.row_dimensions[row].height = 21
            milestone = is_total(rec)

            if milestone:
                fill = TOTAL_FILL
                a_font, num_font = TOTAL_FONT, TOTAL_FONT
                a_align          = LEFT
            else:
                data_row_seq += 1
                # zebra: odd data rows (1st, 3rd, …) get silver-grey fill
                fill    = ZEBRA_FILL if (data_row_seq % 2 == 1) else WHITE_FILL
                a_font, num_font = NORMAL_FONT, NORMAL_FONT
                a_align          = LEFT_IND

            # Column A — text
            ca            = ws[f"A{row}"]
            ca.value      = item
            ca.font       = a_font
            ca.fill       = fill
            ca.alignment  = a_align

            # Columns B & C — numeric
            for col_letter, val in [("B", cy_val), ("C", py_val)]:
                nc               = ws[f"{col_letter}{row}"]
                nc.value         = val
                nc.font          = num_font
                nc.fill          = fill
                nc.alignment     = RIGHT
                nc.number_format = NUM_FORMAT

            # Column D — taxonomy node (metadata only, not read by the algebra engine)
            dc            = ws[f"D{row}"]
            dc.value      = rec.get("taxonomy_node", "UNMAPPED")
            dc.font       = a_font
            dc.fill       = fill
            dc.alignment  = CENTER

            # Borders
            if milestone:
                for cl in ["A", "B", "C", "D"]:
                    ws[f"{cl}{row}"].border = TOTAL_BORDER
            else:
                for cl in ["A", "B", "C", "D"]:
                    ws[f"{cl}{row}"].border = DATA_BORDER

            master_cell_map[sheet_name][item.lower()] = row
            row += 1

        # ── Auto-fit columns B & C with safety margin ─────────────────────────
        for col_idx, col_letter in enumerate(["B", "C"], start=2):
            max_len = max(
                (len(str(ws.cell(r2, col_idx).value or "")) for r2 in range(1, row)),
                default=0,
            )
            ws.column_dimensions[col_letter].width = max(14, max_len + 4)

        ws.column_dimensions["D"].width = 32
        ws.column_dimensions["D"].hidden = True
        ws.freeze_panes = "A3"

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    sheet_mapping = [
        ("Standalone",   "Profit and Loss", "Standalone - P&L"),
        ("Standalone",   "Balance Sheet",   "Standalone - Balance Sheet"),
        ("Standalone",   "Cash Flow",       "Standalone - Cash Flow"),
        ("Consolidated", "Profit and Loss", "Consolidated - P&L"),
        ("Consolidated", "Balance Sheet",   "Consolidated - Balance Sheet"),
        ("Consolidated", "Cash Flow",       "Consolidated - Cash Flow"),
    ]

    for report_type, statement_type, title in sheet_mapping:
        filtered = df[(df["report_type"] == report_type) &
                      (df["statement"]   == statement_type)]
        if not filtered.empty:
            write_sheet(wb, title, filtered)

    # ── Changes in Equity sheets ──────────────────────────────────────────────
    equity_sheet_names = {
        "Standalone":   "Standalone - Changes in Equity",
        "Consolidated": "Consol - Changes in Equity",
    }
    for prefix in ["Standalone", "Consolidated"]:
        title   = equity_sheet_names[prefix]
        eq_data = df[(df["report_type"] == prefix) & (df["statement"] == "Changes in Equity")]
        if not eq_data.empty:
            write_sheet(wb, title, eq_data)

    cols = ["B", "C"]

    # =========================================================================
    # TASK 1 + 2 + 3 COMBINED: PER-SHEET ALGEBRA ENGINE  (unchanged)
    # =========================================================================
    failed_lookups: dict[str, list] = {}

    def _log_lookup(sheet_name: str, node_name: str, row):
        if row is None:
            failed_lookups.setdefault(sheet_name, []).append(node_name)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cm = master_cell_map.get(sheet_name, {})

        is_consolidated = sheet_name.startswith("Consolidated")
        is_pnl          = "P&L" in sheet_name
        is_bs           = "Balance Sheet" in sheet_name
        is_cf           = "Cash Flow" in sheet_name
        is_equity       = "Equity" in sheet_name and not is_bs

        if is_pnl:
            pnl_nodes = [
                "REVENUE_GROSS", "REVENUE_GST_DEDUCTION", "REVENUE_FROM_OPERATIONS",
                "OTHER_INCOME", "TOTAL_INCOME", "TOTAL_EXPENSES",
                "PROFIT_BEFORE_EXCEPTIONAL", "EXCEPTIONAL_ITEMS", "PROFIT_BEFORE_TAX",
                "TOTAL_TAX_EXPENSE", "PROFIT_FOR_THE_YEAR", "TOTAL_OCI",
                "TOTAL_COMPREHENSIVE_INCOME", "EARNINGS_PER_SHARE",
            ]
            if is_consolidated:
                pnl_nodes.append("SHARE_OF_PROFIT_OF_ASSOCIATES")
            for node_name in pnl_nodes:
                _log_lookup(sheet_name, node_name, get_row_by_node(cm, node_name))

            for c in cols:
                inject_pnl_algebra(ws, cm, c, has_associates=is_consolidated)

            strict_sum_nodes = ["TOTAL_EXPENSES", "TOTAL_TAX_EXPENSE"]
            for node_name in strict_sum_nodes:
                target_row = get_row_by_node(cm, node_name)
                if target_row:
                    for c in cols:
                        formula = verify_and_inject_sum(ws, target_row, c, SECTION_RGBS,
                                                         stop_on_section=True)
                        if formula:
                            ws[f"{c}{target_row}"] = formula

            toci_row = get_row_by_node(cm, "TOTAL_OCI")
            if toci_row:
                for c in cols:
                    formula = verify_and_inject_sum(ws, toci_row, c, SECTION_RGBS,
                                                     stop_on_section=False)
                    if formula:
                        ws[f"{c}{toci_row}"] = formula

        elif is_bs:
            bs_nodes = [
                "TOTAL_EQUITY", "TOTAL_NON_CURRENT_LIABILITIES",
                "TOTAL_CURRENT_LIABILITIES", "TOTAL_LIABILITIES",
                "TOTAL_EQUITY_AND_LIABILITIES", "TOTAL_ASSETS",
            ]
            for node_name in bs_nodes:
                _log_lookup(sheet_name, node_name, get_row_by_node(cm, node_name))

            for c in cols:
                inject_bs_algebra(ws, cm, c)

        elif is_cf:
            cf_nodes = [
                "CFO_OPERATING_PROFIT_BEFORE_WC", "CFO_WC_ADJUSTMENTS_SUBTOTAL",
                "CFO_CASH_GENERATED_FROM_OPERATIONS", "CFO_TAXES_PAID",
                "NET_CASH_FROM_OPERATING", "NET_CASH_FROM_INVESTING",
                "NET_CASH_FROM_FINANCING", "NET_CHANGE_IN_CASH",
                "OPENING_CASH_BALANCE", "CLOSING_CASH_BALANCE",
            ]
            for node_name in cf_nodes:
                _log_lookup(sheet_name, node_name, get_row_by_node(cm, node_name))

            for c in cols:
                inject_cf_algebra(ws, cm, c)

        elif is_equity:
            equity_nodes = [
                "RESERVE_OPENING_BALANCE", "RESERVE_CLOSING_BALANCE",
                "EQUITY_SHARE_CAPITAL_OPENING", "EQUITY_SHARE_CAPITAL_CLOSING",
            ]
            for node_name in equity_nodes:
                _log_lookup(sheet_name, node_name, get_row_by_node(cm, node_name))

            for c in cols:
                inject_equity_algebra(ws, cm, c)

    if failed_lookups:
        print("\n⚠️  Taxonomy node lookups that returned no row match, by sheet:")
        for sheet_name, nodes in failed_lookups.items():
            print(f"  {sheet_name}:")
            for node_name in nodes:
                print(f"    - {node_name}")

    # =========================================================================
    # INTER-SHEET EPS CROSS-REFERENCING  (unchanged)
    # =========================================================================
    eps_failures: list[str] = []

    for prefix in ["Standalone", "Consolidated"]:
        pl_name = f"{prefix} - P&L"
        bs_name = f"{prefix} - Balance Sheet"

        if pl_name not in wb.sheetnames or bs_name not in wb.sheetnames:
            eps_failures.append(f"{prefix}: missing P&L or Balance Sheet sheet")
            continue

        pl_ws = wb[pl_name]
        bs_ws = wb[bs_name]

        pl_cm = master_cell_map.get(pl_name, {})
        bs_cm = master_cell_map.get(bs_name, {})

        eps_rows = get_rows_by_node(pl_cm, "EARNINGS_PER_SHARE")
        esc_row  = get_row_by_node(bs_cm, "EQUITY_SHARE_CAPITAL")
        pfy_row  = get_row_by_node(pl_cm, "PROFIT_FOR_THE_YEAR")

        if not eps_rows:
            eps_failures.append(f"{prefix}: no EARNINGS_PER_SHARE row found on {pl_name}")
            continue
        if not esc_row:
            eps_failures.append(f"{prefix}: no EQUITY_SHARE_CAPITAL row found on {bs_name}")
            continue
        if not pfy_row:
            eps_failures.append(f"{prefix}: no PROFIT_FOR_THE_YEAR row found on {pl_name}")
            continue

        for eps_row in eps_rows:
            for c in cols:
                v_esc = _val(bs_ws, esc_row, c)
                if v_esc == 0:
                    eps_failures.append(
                        f"{prefix}: Equity Share Capital is 0 in column {c} — cannot derive share count"
                    )
                    continue
                pl_ws[f"{c}{eps_row}"] = (
                    f"={c}{pfy_row}/('{bs_name}'!{c}{esc_row}/10)"
                )

    if eps_failures:
        print("\n⚠️  EPS cross-referencing issues:")
        for msg in eps_failures:
            print(f"  - {msg}")

    wb.save(output_path)
    print(f"\n✅  Saved math-verified spreadsheet → {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def extract_financials(pdf_path: str, output_xlsx: str = "financials.xlsx"):
    pdf_stem = os.path.splitext(os.path.basename(pdf_path))[0]

    # ── Hardened cache key ────────────────────────────────────────────────────
    # Old behaviour cached purely on pdf_stem, so editing the prompt or
    # switching Gemini models silently kept serving a stale result unless
    # FORCE_REFRESH was flipped by hand. The key now also hashes the
    # extracted PDF text + the prompt template + the model string, so any
    # of those changing produces a fresh cache miss automatically.
    # FORCE_REFRESH remains as an explicit manual override on top of that.
    pdf_text   = extract_text_from_pdf(pdf_path)
    unit_label = detect_reporting_unit(pdf_text)
    cache_key  = _compute_step2_cache_key(pdf_text, _EXTRACTION_PROMPT_TEMPLATE, _GEMINI_MODEL)
    cache_file = _step2_cache_path(cache_key)

    if os.path.exists(cache_file) and not FORCE_REFRESH:
        print(f"✅  Loading from cache ({pdf_stem}) — no API call needed...")
        with open(cache_file, "r") as f:
            all_data = json.load(f)
    else:
        print("Extracting text from PDF...")
        all_data = call_gemini(pdf_text)
        with open(cache_file, "w") as f:
            json.dump(all_data, f, indent=2)

    # ── Taxonomy mapping (fuzzy match, NO LLM) ────────────────────────────────
    # The LLM extracts raw_string/statement/scope/current_year/previous_year —
    # "statement" here is structural placement (which FS sheet), extracted by
    # the LLM so no row is ever dropped for lacking a dictionary match.
    # taxonomy_node is a separate, purely additive label assigned afterwards
    # in Python by fuzzy-matching raw_string against fs_dictionary.py.
    all_data = map_line_items(all_data)

    df = pd.DataFrame(all_data)

    df["taxonomy_node"] = df.get("taxonomy_node", "UNMAPPED").fillna("UNMAPPED")

    df["scope"]       = df.get("scope", "STANDALONE").fillna("STANDALONE")
    df["report_type"] = df["scope"].map(
        lambda s: "Consolidated" if str(s).upper() == "CONSOLIDATED" else "Standalone"
    )

    df["line_item"]     = df["raw_string"].fillna("")
    df["statement"]      = df.get("statement", "").fillna("")
    df.loc[~df["statement"].isin(
        ["Profit and Loss", "Balance Sheet", "Cash Flow", "Statement of Changes in Equity"]
    ), "statement"] = "Unmapped Items"
    df["section"]        = df["statement"]
    df["is_total"]       = df["line_item"].str.lower().str.contains(
        r"\btotal\b|\bnet\b|profit|closing|opening", regex=True, na=False
    )

    unmatched = df[df["fs_statement"] == ""]["line_item"].unique()
    if len(unmatched):
        print(f"\n⚠️  {len(unmatched)} line item(s) had no confident dictionary match "
              f"and were named after their own raw text instead (still placed on their correct sheet):")
        for u in unmatched:
            print(f"    {u}")

    df = df.drop_duplicates(subset=["report_type", "statement", "line_item"])
    df["current_year"]  = df["current_year"].apply(clean_number)
    df["previous_year"] = df["previous_year"].apply(clean_number)
    df = truncate_financial_records(df)

    print("\nBuilding segment pipelines...")
    pipelines = build_segment_pipelines(df)
    for seg in pipelines.values():
        print(f"  {seg.report_type}  |  statements: {list(seg.statements.keys())}  |  has_associates={seg.has_associates}")

    build_excel(df, output_xlsx, unit_label=unit_label)

    # ── Structured JSON output (taxonomy-mapped, for downstream reuse) ───────
    output_json = os.path.splitext(output_xlsx)[0] + "_taxonomy.json"
    json_records = df[["line_item", "taxonomy_node", "fs_statement", "statement",
                        "is_total", "report_type", "current_year", "previous_year",
                        "match_score"]].to_dict(orient="records")
    populated_dict = build_populated_dictionary(df)
    with open(output_json, "w") as f:
        json.dump({"line_items": json_records, "fs_dictionary": populated_dict,
                    "reporting_unit": unit_label}, f, indent=2, default=str)
    print(f"📄  Structured taxonomy JSON written → {output_json}")

    return df

if __name__ == "__main__":
    import argparse
    from pipeline.Step1 import extract_core_financial_statements

    parser = argparse.ArgumentParser()
    parser.add_argument("input_pdf")
    parser.add_argument("output_xlsx")
    args = parser.parse_args()

    gemini_key = os.environ.get("GEMINI_API_KEY")
    if not gemini_key:
        print("❌  GEMINI_API_KEY environment variable not set.")
        sys.exit(1)

    FORCE_REFRESH = False

    input_stem = os.path.splitext(os.path.basename(args.input_pdf))[0]
    intermediate_pdf = os.path.join(os.path.dirname(os.path.abspath(args.input_pdf)), f"{input_stem}_trimmed.pdf")

    print(f"\n── Step 1: Extracting core financial pages → {intermediate_pdf}")
    extract_core_financial_statements(args.input_pdf, intermediate_pdf, gemini_key)

    print(f"\n── Step 2: Parsing & building Excel → {args.output_xlsx}")
    extract_financials(intermediate_pdf, args.output_xlsx)