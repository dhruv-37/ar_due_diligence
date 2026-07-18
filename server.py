"""
server.py
=========
Local Flask server: upload an AR PDF, run the due-diligence pipeline,
render the resulting memo as styled HTML.

Usage:
    python server.py
    -> open http://127.0.0.1:5000
"""
import os, sys, traceback, threading, uuid, io, re
from pathlib import Path
from flask import Flask, request, render_template_string, redirect, url_for, session, jsonify
import markdown as md
import openpyxl

_ROOT = str(Path(__file__).resolve().parent)
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from dotenv import load_dotenv
load_dotenv()

from agents.memo_agent import run_due_diligence
from pipeline.taxonomy import TAXONOMY, Sign, Statement

UPLOAD_DIR = Path(_ROOT) / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "ar-due-diligence-dev-key")

_IMPORTANT_NODES = {
    "REVENUE_FROM_OPERATIONS", "TOTAL_INCOME", "EBITDA", "PROFIT_BEFORE_TAX",
    "PROFIT_FOR_THE_YEAR", "TOTAL_COMPREHENSIVE_INCOME", "EARNINGS_PER_SHARE",
    "TOTAL_ASSETS", "TOTAL_EQUITY", "TOTAL_LIABILITIES", "TOTAL_EQUITY_AND_LIABILITIES",
    "TOTAL_NON_CURRENT_ASSETS", "TOTAL_CURRENT_ASSETS", "TOTAL_NON_CURRENT_LIABILITIES",
    "TOTAL_CURRENT_LIABILITIES", "NET_CASH_FROM_OPERATING", "NET_CASH_FROM_INVESTING",
    "NET_CASH_FROM_FINANCING", "NET_CHANGE_IN_CASH", "CLOSING_CASH_BALANCE",
    "CASH_AND_CASH_EQUIVALENTS", "GROSS_FIXED_ASSETS", "NET_FIXED_ASSETS",
    "NET_WORTH", "GROSS_DEBT", "RESERVE_CLOSING_BALANCE",
}

# ── Row category → background colour (muted, single blue-grey family) ──────
_CATEGORY_COLORS = {
    "heading":    "#dce1f0",   # section banners / subheadings — distinct, cool grey-blue
    "total":      "#c9d6ee",   # subtotals / grand totals — slightly deeper
    "income":     "#e4ecf7",   # P&L inflow items
    "expense":    "#eef1f8",   # P&L outflow items
    "asset":      "#e7edf6",   # balance sheet assets
    "liability":  "#eef0f7",   # balance sheet liabilities / equity
    "cashflow":   "#e9edf6",   # cash flow statement items
    "unmapped":   "#f5f6fa",   # UNMAPPED / unlinked line items
}

def _row_category(node_val: str, is_heading: bool) -> str:
    if is_heading:
        return "heading"
    node = TAXONOMY.get(node_val) if node_val else None
    if node is None:
        return "unmapped"
    if node.is_total:
        return "total"
    if node.statement == Statement.CASH_FLOW:
        return "cashflow"
    if node.statement == Statement.BALANCE_SHEET:
        return "liability" if node.sign == Sign.NEGATIVE else "asset"
    # Profit and Loss / Changes in Equity
    return "expense" if node.sign == Sign.NEGATIVE else "income"


# ── Background job tracking ─────────────────────────────────────────────────
_JOBS = {}
_JOBS_LOCK = threading.Lock()

_STAGES = ["EXTRACT", "RED FLAGS", "NARRATIVE", "WRITE MEMO"]
_STAGE_MARKER = re.compile(r"══ Node:\s*(.+?)\s*═+")


class _JobStream(io.TextIOBase):
    """Captures stdout line-by-line into a job's live log."""
    def __init__(self, job_id):
        self.job_id = job_id
        self._buf = ""

    def write(self, s):
        self._buf += s
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            if line.strip():
                _push_line(self.job_id, line.strip())
        return len(s)

    def flush(self):
        pass


def _push_line(job_id, line):
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        if not job:
            return
        job["lines"].append(line)
        m = _STAGE_MARKER.search(line)
        if m:
            stage_name = m.group(1).strip().upper()
            for idx, s in enumerate(_STAGES):
                if s in stage_name:
                    job["stage_idx"] = idx
                    break


def _run_job(job_id, pdf_path, ticker):
    stream = _JobStream(job_id)
    old_stdout = sys.stdout
    sys.stdout = stream
    try:
        result = run_due_diligence(pdf_path=pdf_path, ticker=ticker)
        memo_text = result.get("memo_text", "# No memo generated")
        memo_html = md.markdown(memo_text, extensions=["extra", "tables", "sane_lists"])
        xlsx_path = result.get("extraction", {}).get("output_xlsx", "")
        with _JOBS_LOCK:
            _JOBS[job_id]["memo_html"] = memo_html
            _JOBS[job_id]["xlsx_path"] = xlsx_path
            _JOBS[job_id]["stage_idx"] = len(_STAGES) - 1
            _JOBS[job_id]["done"] = True
    except Exception as exc:
        traceback.print_exc()
        with _JOBS_LOCK:
            _JOBS[job_id]["error"] = str(exc)
            _JOBS[job_id]["done"] = True
    finally:
        sys.stdout = old_stdout

PAGE = """
<!doctype html>
<html><head>
<meta charset="utf-8">
<title>AR Due Diligence</title>
<style>
  body { font-family: 'Segoe UI', sans-serif; max-width: 860px; margin: 40px auto; padding: 0 20px; background: #f7f8fb; color: #1f2430; }
  h1 { color: #2b2f77; }
  form { background: #fff; padding: 24px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,.08); }
  input[type=text], input[type=file] { padding: 8px; margin: 8px 0 16px; width: 100%; box-sizing: border-box; border: 1px solid #ccc; border-radius: 6px; }
  button { background: #4f5bd5; color: #fff; border: none; padding: 10px 22px; border-radius: 8px; font-size: 15px; cursor: pointer; }
  button:hover { background: #3c46b0; }
  .memo { background: #fff; padding: 32px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,.08); margin-top: 24px; }
  .memo h1, .memo h2, .memo h3 { color: #2b2f77; }
  .memo table { border-collapse: collapse; width: 100%; margin: 12px 0; }
  .memo th, .memo td { border: 1px solid #ddd; padding: 8px 10px; text-align: left; }
  .memo th { background: #eef0fb; }
  .memo code { background: #f0f0f5; padding: 2px 5px; border-radius: 4px; }
  .memo pre { background: #f0f0f5; padding: 12px; border-radius: 8px; overflow-x: auto; }
  .error { background: #ffe9e9; color: #a30000; padding: 16px; border-radius: 8px; }
  .flags li { margin-bottom: 6px; }
  .actions { margin-top: 16px; display:flex; gap:10px; }
  .btn-secondary { background:#00b39f; color:#fff; border:none; padding:10px 22px; border-radius:8px; font-size:15px; cursor:pointer; text-decoration:none; display:inline-block; }
  .btn-secondary:hover { background:#00937f; }
  .tabs { display:flex; gap:6px; flex-wrap:wrap; margin-bottom:14px; }
  .tab-btn { background:#eef0fb; border:none; padding:8px 16px; border-radius:20px; cursor:pointer; font-weight:600; color:#2b2f77; }
  .tab-btn.active { background:#4f5bd5; color:#fff; }
  .sheet-table { display:none; overflow-x:auto; }
  .sheet-table.active { display:block; }
  table.xl { border-collapse:collapse; width:100%; font-size:13px; }
  table.xl th { background:#4f5bd5; color:#fff; padding:8px 10px; position:sticky; top:0; }
  table.xl td { padding:6px 10px; border:1px solid #e2e4f0; }
  table.xl tr:nth-child(even) td { background:#f4f5fc; }
  table.xl tr:hover td { background:#e9ecfc; }
  .num { text-align:right; color:#1a3a8f; font-variant-numeric: tabular-nums; }
</style>
<script>
function showSheet(name){
  document.querySelectorAll('.sheet-table').forEach(el=>el.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(el=>el.classList.remove('active'));
  document.getElementById('sheet-'+name).classList.add('active');
  document.getElementById('tab-'+name).classList.add('active');
}
</script>
</head><body>
<h1>📊 AR Due Diligence Agent</h1>
<form method="post" action="{{ url_for('start') }}" enctype="multipart/form-data">
  <label>Annual Report PDF</label>
  <input type="file" name="pdf" accept="application/pdf" required>
  <label>Ticker (NSE, e.g. RELIANCE)</label>
  <input type="text" name="ticker" required>
  <button type="submit">Run Due Diligence</button>
</form>
{% if error %}
<div class="error"><b>Error:</b> {{ error }}</div>
{% endif %}
{% if memo_html %}
<div class="actions">
  {% if xlsx_path %}
  <a class="btn-secondary" href="{{ url_for('view_excel', path=xlsx_path) }}">📈 View Excel Output</a>
  {% endif %}
  <a class="btn-secondary" style="background:#888" href="{{ url_for('index', reset=1) }}">🔄 New Analysis</a>
</div>
<div class="memo">{{ memo_html | safe }}</div>
{% endif %}
</body></html>
"""

EXCEL_PAGE = """
<!doctype html>
<html><head>
<meta charset="utf-8">
<title>Excel Output</title>
<style>
  body { font-family: 'Segoe UI', sans-serif; max-width: 1100px; margin: 40px auto; padding: 0 20px; background: #f7f8fb; color: #1f2430; }
  h1 { color: #2b2f77; }
  a.back { color:#4f5bd5; text-decoration:none; font-weight:600; }
  .card { background:#fff; padding:24px; border-radius:12px; box-shadow:0 2px 10px rgba(0,0,0,.08); margin-top:16px; }
  .tabs { display:flex; gap:6px; flex-wrap:wrap; margin-bottom:14px; }
  .tab-btn { background:#eef0fb; border:none; padding:8px 16px; border-radius:20px; cursor:pointer; font-weight:600; color:#2b2f77; }
  .tab-btn.active { background:#4f5bd5; color:#fff; }
  .sheet-table { display:none; overflow-x:auto; }
  .sheet-table.active { display:block; }
  table.xl { border-collapse:collapse; width:100%; font-size:13px; }
  table.xl th { background:#4f5bd5; color:#fff; padding:8px 10px; position:sticky; top:0; }
  table.xl td { padding:6px 10px; border:1px solid #e2e4f0; }
  table.xl tr:nth-child(even) td { background:#f4f5fc; }
  table.xl tr:hover td { background:#e9ecfc; }
  .num { text-align:right; color:#1a3a8f; font-variant-numeric: tabular-nums; }
  .important-row td { background:#fff6da !important; font-weight:700; }
  .important-row td.num.important { color:#b8860b; }
  td.important:not(.num) { color:#1a3a1a; }
</style>
<script>
function showSheet(name){
  document.querySelectorAll('.sheet-table').forEach(el=>el.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(el=>el.classList.remove('active'));
  document.getElementById('sheet-'+name).classList.add('active');
  document.getElementById('tab-'+name).classList.add('active');
}
</script>
</head><body>
<a class="back" href="{{ url_for('index') }}">&larr; Back to Memo</a>
<h1>📈 Excel Output</h1>
<div class="card">
  <div class="tabs">
    {% for name in sheet_names %}
    <button class="tab-btn {{ 'active' if loop.first else '' }}" id="tab-{{ name|replace(' ','_') }}" onclick="showSheet('{{ name|replace(' ','_') }}')">{{ name }}</button>
    {% endfor %}
  </div>
  {% for name, table_html in sheets %}
  <div class="sheet-table {{ 'active' if loop.first else '' }}" id="sheet-{{ name|replace(' ','_') }}">{{ table_html | safe }}</div>
  {% endfor %}
</div>
</body></html>
"""

PROGRESS_PAGE = """
<!doctype html>
<html><head>
<meta charset="utf-8">
<title>Running Due Diligence…</title>
<style>
  body { font-family: 'Segoe UI', sans-serif; max-width: 780px; margin: 40px auto; padding: 0 20px; background: #0e1020; color: #d9e0ff; }
  h1 { color: #8fa2ff; text-align:center; }
  .stages { display:flex; justify-content:space-between; margin: 28px 0 18px; }
  .stage { flex:1; text-align:center; position:relative; }
  .stage .dot { width:34px; height:34px; border-radius:50%; background:#262a4d; border:2px solid #444a80; margin:0 auto 8px; display:flex; align-items:center; justify-content:center; font-weight:700; transition: all .4s ease; }
  .stage.active .dot { background:#4f5bd5; border-color:#8fa2ff; box-shadow:0 0 18px #4f5bd5; animation: pulse 1.2s infinite; }
  .stage.done .dot { background:#00b39f; border-color:#00e0c0; box-shadow:0 0 12px #00b39f; }
  .stage span.label { font-size:12px; color:#9aa4d6; }
  .stage.active span.label { color:#c9d2ff; font-weight:700; }
  .bar-track { height:6px; background:#20223e; border-radius:6px; overflow:hidden; margin-bottom:26px; }
  .bar-fill { height:100%; width:0%; background:linear-gradient(90deg,#4f5bd5,#00b39f); transition: width .6s ease; }
  @keyframes pulse { 0%{ transform:scale(1);} 50%{ transform:scale(1.15);} 100%{ transform:scale(1);} }
  .terminal { background:#05060f; border:1px solid #262a4d; border-radius:10px; padding:18px 20px; height:340px; overflow-y:auto; font-family:'Consolas','Menlo',monospace; font-size:13px; line-height:1.6; box-shadow: inset 0 0 30px rgba(79,91,213,.08); }
  .terminal .ln { opacity:0; animation: fadeIn .35s forwards; white-space:pre-wrap; }
  .terminal .ln.ok { color:#5be08a; }
  .terminal .ln.warn { color:#ffcf5c; }
  .terminal .ln.node { color:#8fa2ff; font-weight:700; margin-top:8px; }
  .terminal .ln.plain { color:#9aa4d6; }
  @keyframes fadeIn { from{opacity:0; transform:translateY(4px);} to{opacity:1; transform:translateY(0);} }
  .cursor { display:inline-block; width:8px; height:14px; background:#8fa2ff; animation: blink 1s step-end infinite; vertical-align:middle; margin-left:2px; }
  @keyframes blink { 50% { opacity:0; } }
  .errbox { background:#3a0d16; color:#ff8f8f; padding:14px 18px; border-radius:8px; margin-top:16px; }
</style>
</head><body>
<h1>⚙️ Running Due Diligence…</h1>
<div class="stages" id="stages">
  {% for s in stages %}
  <div class="stage" id="stage-{{ loop.index0 }}">
    <div class="dot">{{ loop.index }}</div>
    <span class="label">{{ s }}</span>
  </div>
  {% endfor %}
</div>
<div class="bar-track"><div class="bar-fill" id="barFill"></div></div>
<div class="terminal" id="terminal"></div>
<div id="errWrap"></div>

<script>
const jobId = "{{ job_id }}";
const totalStages = {{ stages|length }};
let shownLines = 0;

function classify(line){
  if (line.includes('══ Node:')) return 'node';
  if (line.startsWith('  ✅') || line.includes('✅')) return 'ok';
  if (line.startsWith('  ⚠️') || line.includes('⚠️')) return 'warn';
  return 'plain';
}

function updateStages(stageIdx, done){
  for (let i = 0; i < totalStages; i++){
    const el = document.getElementById('stage-' + i);
    el.classList.remove('active','done');
    if (i < stageIdx || (done && i <= stageIdx)) el.classList.add('done');
    else if (i === stageIdx) el.classList.add('active');
  }
  const pct = done ? 100 : Math.min(95, ((stageIdx + 0.5) / totalStages) * 100);
  document.getElementById('barFill').style.width = pct + '%';
}

async function poll(){
  try {
    const res = await fetch('/status/' + jobId);
    const data = await res.json();
    const term = document.getElementById('terminal');
    for (; shownLines < data.lines.length; shownLines++){
      const div = document.createElement('div');
      div.className = 'ln ' + classify(data.lines[shownLines]);
      div.textContent = data.lines[shownLines];
      term.appendChild(div);
    }
    term.scrollTop = term.scrollHeight;
    updateStages(data.stage_idx, data.done);

    if (data.error){
      document.getElementById('errWrap').innerHTML = '<div class="errbox"><b>Error:</b> ' + data.error + '</div>';
      return;
    }
    if (data.done){
      setTimeout(() => { window.location.href = '/result/' + jobId; }, 700);
      return;
    }
    setTimeout(poll, 600);
  } catch(e){
    setTimeout(poll, 1200);
  }
}
poll();
</script>
</body></html>
"""

@app.route("/", methods=["GET"])
def index():
    if request.args.get("reset"):
        session.pop("memo_html", None)
        session.pop("xlsx_path", None)
        return redirect(url_for("index"))
    return render_template_string(
        PAGE,
        memo_html=session.get("memo_html"),
        error=None,
        xlsx_path=session.get("xlsx_path"),
    )


@app.route("/start", methods=["POST"])
def start():
    f = request.files.get("pdf")
    ticker = request.form.get("ticker", "").strip().upper()
    if not f or not f.filename.lower().endswith(".pdf") or not ticker:
        return render_template_string(PAGE, memo_html=None, error="Please provide a valid PDF and ticker.", xlsx_path=None)

    pdf_path = UPLOAD_DIR / f.filename
    f.save(pdf_path)

    job_id = uuid.uuid4().hex
    with _JOBS_LOCK:
        _JOBS[job_id] = {"lines": [], "stage_idx": 0, "done": False, "error": None,
                          "memo_html": None, "xlsx_path": None}

    t = threading.Thread(target=_run_job, args=(job_id, str(pdf_path), ticker), daemon=True)
    t.start()

    return redirect(url_for("progress_page", job_id=job_id))


@app.route("/progress/<job_id>")
def progress_page(job_id):
    if job_id not in _JOBS:
        return redirect(url_for("index"))
    return render_template_string(PROGRESS_PAGE, job_id=job_id, stages=_STAGES)


@app.route("/status/<job_id>")
def status(job_id):
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        if not job:
            return jsonify({"lines": [], "stage_idx": 0, "done": True, "error": "Job not found"})
        return jsonify({
            "lines": job["lines"],
            "stage_idx": job["stage_idx"],
            "done": job["done"],
            "error": job["error"],
        })


@app.route("/result/<job_id>")
def result(job_id):
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
    if not job:
        return redirect(url_for("index"))
    session["memo_html"] = job.get("memo_html")
    session["xlsx_path"] = job.get("xlsx_path")
    with _JOBS_LOCK:
        _JOBS.pop(job_id, None)
    return redirect(url_for("index"))


def _cell_html(value, bold=False, bg=None):
    cls = "num" if isinstance(value, (int, float)) else ""
    cls = (cls + " important").strip() if bold else cls
    cls_attr = f' class="{cls}"' if cls else ""
    style_attr = f' style="background:{bg};"' if bg else ""
    if isinstance(value, float):
        return f'<td{cls_attr}{style_attr}>{value:,.2f}</td>'
    if isinstance(value, int):
        return f'<td{cls_attr}{style_attr}>{value:,}</td>'
    return f"<td{cls_attr}{style_attr}>{'' if value is None else value}</td>"


@app.route("/excel")
def view_excel():
    path = request.args.get("path", "")
    if not path or not Path(path).exists():
        return render_template_string(PAGE, memo_html=session.get("memo_html"), error="Excel file not found.", xlsx_path=None)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames
    sheets = []
    for name in sheet_names:
        ws = wb[name]
        rows_html = []
        header = None
        taxonomy_col = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row = list(row)
            if i == 0:
                header = row
                taxonomy_col = next(
                    (idx for idx, h in enumerate(header) if h and "taxonomy" in str(h).lower()),
                    None,
                )
                display_row = [v for idx, v in enumerate(row) if idx != taxonomy_col]
                cells = "".join(f"<th>{'' if v is None else v}</th>" for v in display_row)
                rows_html.append(f"<tr>{cells}</tr>")
                continue

            node_val = str(row[taxonomy_col]).strip().upper() if taxonomy_col is not None and taxonomy_col < len(row) else ""
            is_important = node_val in _IMPORTANT_NODES
            display_row = [v for idx, v in enumerate(row) if idx != taxonomy_col]
            # Heading / subheading rows: no numeric values in the FY columns
            numeric_vals = [v for idx, v in enumerate(row)
                             if idx != taxonomy_col and idx != 0
                             and isinstance(v, (int, float))]
            is_heading = len(numeric_vals) == 0
            category = _row_category(node_val, is_heading)
            row_color = _CATEGORY_COLORS[category]
            cells = "".join(_cell_html(v, bold=is_important, bg=None if is_important else row_color) for v in display_row)
            row_cls = ' class="important-row"' if is_important else ""
            rows_html.append(f"<tr{row_cls}>{cells}</tr>")

        table_html = f'<table class="xl">{"".join(rows_html)}</table>'
        sheets.append((name, table_html))

    return render_template_string(EXCEL_PAGE, sheet_names=sheet_names, sheets=sheets)


if __name__ == "__main__":
    missing = [k for k in ["GEMINI_API_KEY"] if not os.environ.get(k)]
    if missing:
        print(f"⚠️  Missing env vars: {missing} — add to .env before running.")
    app.run(host="127.0.0.1", port=5000, debug=True)