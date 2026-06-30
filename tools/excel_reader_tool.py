"""
tools/excel_reader_tool.py
==========================
Reads the structured Excel file produced by Step2 into a JSON-serialisable
dict so downstream agents (Red Flag, Narrative, Memo) can consume it.

Each sheet becomes a list of {line_item, current_year, previous_year,
taxonomy_node} records. Empty or missing sheets are skipped silently.
"""

import json
import os
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from langchain.tools import tool

_SEGMENT_PATTERNS = [("standalone", r"\bstand[\s\-_]*alone\b"), ("consolidated", r"\bconsol(idated)?\b")]
_STATEMENT_PATTERNS = [
    ("balance_sheet", r"\bbalance[\s\-_]*sheet\b|\bbs\b"),
    ("pnl",           r"\bp[\s\-_]*&?[\s\-_]*l\b|\bprofit[\s\-_]*(and|&)?[\s\-_]*loss\b"),
    ("cash_flow",     r"\bcash[\s\-_]*flow\b|\bcf\b"),
    ("equity",        r"\bequity\b|\bchanges?\s+in\s+equity\b"),
]

# Matches a cell reference, optionally prefixed by a quoted cross-sheet name,
# e.g. "B21", "$B$21", or "'Standalone - Balance Sheet'!B10".
_CELL_REF_RE = re.compile(r"(?:'(?P<sheet>[^']+)'!)?(?P<ref>\$?[A-Z]{1,3}\$?\d+)")


def _resolve_formula_cell(wb_values, wb_formulas, sheet_name: str, cell_ref: str,
                           memo: dict, visiting: set):
    """
    Resolves a single cell's value, evaluating Excel formulas in Python when
    the cell has no cached calculated value (i.e. the workbook was built and
    saved by openpyxl/Step2 and never opened in real Excel, so formula
    results were never cached).

    Handles the simple +/-/() arithmetic chains and the cross-sheet division
    formulas that Step2's algebra-injection functions write. Returns None if
    the cell is blank, unresolvable, or part of a circular reference.
    """
    key = (sheet_name, cell_ref)
    if key in memo:
        return memo[key]
    if key in visiting or sheet_name not in wb_values.sheetnames:
        return None
    visiting.add(key)

    cached = wb_values[sheet_name][cell_ref].value
    if isinstance(cached, (int, float)) and cached == cached:  # not NaN
        memo[key] = cached
        visiting.discard(key)
        return cached

    formula = wb_formulas[sheet_name][cell_ref].value if sheet_name in wb_formulas.sheetnames else None
    value = None
    if isinstance(formula, str) and formula.startswith("="):
        def _replace(m: "re.Match") -> str:
            target_sheet = m.group("sheet") or sheet_name
            ref = m.group("ref").replace("$", "")
            resolved = _resolve_formula_cell(wb_values, wb_formulas, target_sheet, ref, memo, visiting)
            return repr(float(resolved)) if resolved is not None else "0"

        expr = _CELL_REF_RE.sub(_replace, formula[1:])
        try:
            # Expression now contains only numbers/operators/parens — safe to eval.
            value = eval(expr, {"__builtins__": {}}, {})
        except Exception:
            value = None

    visiting.discard(key)
    memo[key] = value
    return value


def _canonical_sheet_name(sheet_name: str) -> str:
    """
    Maps loosely-formatted sheet names (e.g. 'Standalone - Balance Sheet',
    'CONSOLIDATED_BS', 'Standalone P&L') to a canonical
    '<segment>_<statement>' key. Falls back to the original name if no
    segment/statement pattern is recognised.
    """
    s = str(sheet_name).strip().lower()
    segment = next((seg for seg, pat in _SEGMENT_PATTERNS if re.search(pat, s)), None)
    statement = next((st for st, pat in _STATEMENT_PATTERNS if re.search(pat, s)), None)
    if segment and statement:
        return f"{segment}_{statement}"
    return sheet_name


@tool
def excel_reader_tool(xlsx_path: str) -> str:
    """
    Reads the structured financial Excel file produced by Step2.

    Args:
        xlsx_path: Absolute or relative path to the .xlsx file.

    Returns:
        JSON string with keys:
            status  — "success" | "error"
            sheets  — dict of sheet_name → list of row dicts
            error   — error message (only present on failure)
    """
    xlsx_path = str(Path(xlsx_path).resolve())

    if not Path(xlsx_path).exists():
        return json.dumps({"status": "error", "error": f"File not found: {xlsx_path}"})

    try:
        xf = pd.ExcelFile(xlsx_path)
        # Loaded once, reused across all sheets, for resolving formula cells
        # that Step2's algebra-injection left without a cached calculated
        # value (the workbook is built/saved by openpyxl, never opened in
        # real Excel, so formula results are never cached there).
        wb_values   = load_workbook(xlsx_path, data_only=True)
        wb_formulas = load_workbook(xlsx_path, data_only=False)
        formula_memo: dict = {}

        sheets = {}

        for sheet_name in xf.sheet_names:
            df = pd.read_excel(xf, sheet_name=sheet_name)
            header_row_used = 0

            # ── Handle sheets where row 0 is a title (e.g. 'Standalone - P&L')
            #    rather than the real header, pushing 'Line Item' / 'FYxx' down
            #    into the first data row. Detect this and re-read with header=1.
            def _has_header_keywords(cols) -> bool:
                low = [str(c).strip().lower() for c in cols]
                return any("line item" in c or "line_item" in c for c in low) and \
                       any("fy" in c or "20" in c for c in low)

            if not _has_header_keywords(df.columns):
                df_retry = pd.read_excel(xf, sheet_name=sheet_name, header=1)
                if _has_header_keywords(df_retry.columns):
                    df = df_retry
                    header_row_used = 1

            # ── Smart Column Mapping ──────────────────────────────────────────
            col_map = {}
            year_cols = []
            
            for c in df.columns:
                c_lower = str(c).strip().lower()
                
                if "line item" in c_lower or "line_item" in c_lower:
                    col_map[c] = "line_item"
                elif "taxonomy" in c_lower:
                    col_map[c] = "taxonomy_node"
                elif "fy" in c_lower or "20" in c_lower or "₹" in c_lower or "lakhs" in c_lower:
                    # Catch financial year columns (e.g., 'FY25 (₹ Lakhs)')
                    year_cols.append(c)

            # Indian ARs traditionally list Current Year first, then Previous Year
            if len(year_cols) >= 1:
                col_map[year_cols[0]] = "current_year"
            if len(year_cols) >= 2:
                col_map[year_cols[1]] = "previous_year"

            # ── Capture original column positions BEFORE renaming, so blank
            #    cells can be mapped back to an Excel address (e.g. "B21")
            #    for formula resolution further down.
            orig_columns = list(df.columns)
            current_year_letter  = (get_column_letter(orig_columns.index(year_cols[0]) + 1)
                                     if len(year_cols) >= 1 else None)
            previous_year_letter = (get_column_letter(orig_columns.index(year_cols[1]) + 1)
                                     if len(year_cols) >= 2 else None)
            # Excel is 1-indexed; header occupies row (header_row_used + 1),
            # so data rows start the row after that.
            data_start_excel_row = header_row_used + 2

            df.rename(columns=col_map, inplace=True)
            # ──────────────────────────────────────────────────────────────────

            # Keep only the four core columns; fill missing ones with None
            core_cols = ["line_item", "current_year", "previous_year", "taxonomy_node"]
            for col in core_cols:
                if col not in df.columns:
                    df[col] = None

            df = df[core_cols].dropna(subset=["line_item"])

            if df.empty:
                continue

            # ── Resolve blank current_year/previous_year cells that are
            #    actually unevaluated Excel formulas (see _resolve_formula_cell).
            for idx, row in df.iterrows():
                excel_row = data_start_excel_row + idx
                if pd.isna(row["current_year"]) and current_year_letter:
                    resolved = _resolve_formula_cell(
                        wb_values, wb_formulas, sheet_name,
                        f"{current_year_letter}{excel_row}", formula_memo, set())
                    if resolved is not None:
                        df.at[idx, "current_year"] = resolved
                if pd.isna(row["previous_year"]) and previous_year_letter:
                    resolved = _resolve_formula_cell(
                        wb_values, wb_formulas, sheet_name,
                        f"{previous_year_letter}{excel_row}", formula_memo, set())
                    if resolved is not None:
                        df.at[idx, "previous_year"] = resolved

            # Convert NaN → None for JSON serialisation
            records = df.where(pd.notna(df), other=None).to_dict(orient="records")
            sheets[_canonical_sheet_name(sheet_name)] = records

        if not sheets:
            return json.dumps({"status": "error", "error": "No usable sheets found in Excel file."})

        return json.dumps({"status": "success", "sheets": sheets})

    except Exception as exc:
        return json.dumps({"status": "error", "error": str(exc)})